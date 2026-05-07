"""
XLSX'i 50 satırdan 200 satıra genişletir.
Tüm formülleri, data validasyonları ve referansları günceller.
"""
import os
import sys
import shutil
from datetime import datetime

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'autofix_qa_tracker.xlsx')
OLD_MAX_ROW = 54   # Eski son satır
NEW_MAX_ROW = 204  # Yeni son satır (200 veri satırı)
DATA_START = 5     # Veri başlangıç satırı


def expand():
    # Yedek
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = XLSX_PATH.replace('.xlsx', f'_{ts}_pre_expand.xlsx')
    shutil.copy2(XLSX_PATH, backup)
    print(f"✅ Yedek: {backup}")

    wb = openpyxl.load_workbook(XLSX_PATH)

    # ═══════════════════════════════════════
    # 1. TEST LOG — Genişlet
    # ═══════════════════════════════════════
    ws = wb['Test Log']

    # Önce eski dry-run verisini temizle
    for row in range(DATA_START, OLD_MAX_ROW + 1):
        for col in range(2, 19):  # B-R sütunları (A ve S formül)
            ws.cell(row=row, column=col).value = None

    # Eski data validasyonlarını kaldır
    ws.data_validations.dataValidation.clear()

    # Yeni satırlar için formülleri oluştur
    for row in range(DATA_START, NEW_MAX_ROW + 1):
        # A: Test ID
        ws.cell(row=row, column=1).value = f'=IF(B{row}="","",ROW()-4)'
        # E: Senaryo Adı (VLOOKUP)
        ws.cell(row=row, column=5).value = f"=IFERROR(VLOOKUP(D{row},'Senaryo Analizi'!A$5:B$19,2,FALSE),\"\")"
        # F: Araç/Arıza (VLOOKUP)
        ws.cell(row=row, column=6).value = f"=IFERROR(VLOOKUP(D{row},'Senaryo Analizi'!A$5:C$19,3,FALSE),\"\")"
        # G: Zorluk (otomatik)
        ws.cell(row=row, column=7).value = f'=IF(D{row}="","",IF(VALUE(MID(D{row},2,2))<=5,"Kolay",IF(VALUE(MID(D{row},2,2))<=10,"Orta","Zor")))'
        # S: Flag
        ws.cell(row=row, column=19).value = f'=IF(OR(K{row}="Evet",N{row}="Evet",H{row}="FAIL"),"🚩","")'

    # Yeni data validasyonları (genişletilmiş aralık)
    rng = f"D{DATA_START}:D{NEW_MAX_ROW}"
    dv1 = DataValidation(type="list", formula1='"S01,S02,S03,S04,S05,S06,S07,S08,S09,S10,S11,S12,S13,S14,S15"', allow_blank=True)
    dv1.sqref = rng
    ws.add_data_validation(dv1)

    dv2 = DataValidation(type="list", formula1='"PASS,FAIL,PARTIAL"', allow_blank=True)
    dv2.sqref = f"H{DATA_START}:H{NEW_MAX_ROW}"
    ws.add_data_validation(dv2)

    dv3 = DataValidation(type="list", formula1='"Evet,Hayır"', allow_blank=True)
    dv3.sqref = f"J{DATA_START}:K{NEW_MAX_ROW} N{DATA_START}:O{NEW_MAX_ROW}"
    ws.add_data_validation(dv3)

    dv4 = DataValidation(type="list", formula1='"Yok,Uydurma Parça,Yanlış Teşhis,Protokol Dışı,Dil Karışımı,Diğer"', allow_blank=True)
    dv4.sqref = f"L{DATA_START}:L{NEW_MAX_ROW}"
    ws.add_data_validation(dv4)

    dv5 = DataValidation(type="list", formula1='"Kolay,Orta,Zor"', allow_blank=True)
    dv5.sqref = f"G{DATA_START}:G{NEW_MAX_ROW}"
    ws.add_data_validation(dv5)

    print(f"✅ Test Log: {NEW_MAX_ROW - DATA_START + 1} satıra genişletildi")

    # ═══════════════════════════════════════
    # 2. SENARYO ANALİZİ — Formül referanslarını güncelle
    # ═══════════════════════════════════════
    ws2 = wb['Senaryo Analizi']
    for row in range(5, 20):  # 15 senaryo satırı
        ref = f"A{row}"
        # D: Test Sayısı
        ws2.cell(row=row, column=4).value = f"=COUNTIF('Test Log'!D${DATA_START}:D${NEW_MAX_ROW},{ref})"
        # E: PASS
        ws2.cell(row=row, column=5).value = f'=COUNTIFS(\'Test Log\'!D${DATA_START}:D${NEW_MAX_ROW},{ref},\'Test Log\'!H${DATA_START}:H${NEW_MAX_ROW},"PASS")'
        # F: FAIL
        ws2.cell(row=row, column=6).value = f'=COUNTIFS(\'Test Log\'!D${DATA_START}:D${NEW_MAX_ROW},{ref},\'Test Log\'!H${DATA_START}:H${NEW_MAX_ROW},"FAIL")'
        # G: PARTIAL
        ws2.cell(row=row, column=7).value = f'=COUNTIFS(\'Test Log\'!D${DATA_START}:D${NEW_MAX_ROW},{ref},\'Test Log\'!H${DATA_START}:H${NEW_MAX_ROW},"PARTIAL")'
        # H: Başarı Oranı %
        ws2.cell(row=row, column=8).value = f"=IFERROR(E{row}/D{row},0)"
        # I: Ortalama Mesaj Sayısı
        ws2.cell(row=row, column=9).value = f"=IFERROR(AVERAGEIF('Test Log'!D${DATA_START}:D${NEW_MAX_ROW},{ref},'Test Log'!I${DATA_START}:I${NEW_MAX_ROW}),0)"
        # J: Halüsinasyon Sayısı
        ws2.cell(row=row, column=10).value = f'=COUNTIFS(\'Test Log\'!D${DATA_START}:D${NEW_MAX_ROW},{ref},\'Test Log\'!K${DATA_START}:K${NEW_MAX_ROW},"Evet")'
        # K: Hal. Oranı %
        ws2.cell(row=row, column=11).value = f"=IFERROR(J{row}/D{row},0)"
        # L: Risk Seviyesi
        ws2.cell(row=row, column=12).value = f'=IF(K{row}>0.3,"🔴 Yüksek",IF(K{row}>0.1,"🟡 Orta","🟢 Düşük"))'

    print(f"✅ Senaryo Analizi: Formüller D$5:D$204 referansına güncellendi")

    # ═══════════════════════════════════════
    # 3. DASHBOARD — Formülleri güncelle
    # ═══════════════════════════════════════
    ws3 = wb['Dashboard']
    # Genel istatistikler (Row 5)
    ws3['B5'].value = f"=COUNTA('Test Log'!B{DATA_START}:B{NEW_MAX_ROW})"
    ws3['D5'].value = f'=COUNTIF(\'Test Log\'!H{DATA_START}:H{NEW_MAX_ROW},"PASS")'
    ws3['F5'].value = f'=COUNTIF(\'Test Log\'!H{DATA_START}:H{NEW_MAX_ROW},"FAIL")'
    ws3['H5'].value = f'=COUNTIF(\'Test Log\'!H{DATA_START}:H{NEW_MAX_ROW},"PARTIAL")'
    ws3['J5'].value = f'=IFERROR(COUNTIF(\'Test Log\'!H{DATA_START}:H{NEW_MAX_ROW},"PASS")/COUNTA(\'Test Log\'!H{DATA_START}:H{NEW_MAX_ROW}),0)'
    ws3['L5'].value = f'=COUNTIF(\'Test Log\'!K{DATA_START}:K{NEW_MAX_ROW},"Evet")'
    ws3['N5'].value = f'=IFERROR(COUNTIF(\'Test Log\'!K{DATA_START}:K{NEW_MAX_ROW},"Evet")/COUNTA(\'Test Log\'!B{DATA_START}:B{NEW_MAX_ROW}),0)'

    # Halüsinasyon Tip Dağılımı (K9-K13)
    for row in range(9, 14):
        ws3.cell(row=row, column=11).value = f"=COUNTIF('Test Log'!L${DATA_START}:L${NEW_MAX_ROW},J{row})"

    print(f"✅ Dashboard: Formüller güncellendi")

    # ═══════════════════════════════════════
    # 4. HALÜSİNASYON DETAY — Genişlet
    # ═══════════════════════════════════════
    ws4 = wb['Halüsinasyon Detay']
    # Eski validasyonları kaldır
    ws4.data_validations.dataValidation.clear()

    # Yeni validasyonlar
    dv_tip = DataValidation(type="list", formula1='"Yok,Uydurma Parça,Yanlış Teşhis,Protokol Dışı,Dil Karışımı,Diğer"', allow_blank=True)
    dv_tip.sqref = f"E{DATA_START}:E{NEW_MAX_ROW}"
    ws4.add_data_validation(dv_tip)

    dv_katman = DataValidation(type="list", formula1='"Giriş Filtresi,Prompt Hardening,Çıkış Filtresi,Tespit Edilemedi"', allow_blank=True)
    dv_katman.sqref = f"I{DATA_START}:I{NEW_MAX_ROW}"
    ws4.add_data_validation(dv_katman)

    dv_aksiyon = DataValidation(type="list", formula1='"İnceleniyor,Prompt Güncellendi,Lambda Güncellendi,Kapatıldı,Yok Sayıldı"', allow_blank=True)
    dv_aksiyon.sqref = f"J{DATA_START}:J{NEW_MAX_ROW}"
    ws4.add_data_validation(dv_aksiyon)

    print(f"✅ Halüsinasyon Detay: {NEW_MAX_ROW - DATA_START + 1} satıra genişletildi")

    # ═══════════════════════════════════════
    # 5. Kaydet
    # ═══════════════════════════════════════
    wb.save(XLSX_PATH)
    print(f"\n🎉 XLSX başarıyla genişletildi! Yeni kapasite: {NEW_MAX_ROW - DATA_START + 1} test satırı")


if __name__ == '__main__':
    expand()
