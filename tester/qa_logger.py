"""
Auto Fix AI — QA Test Kayıt Scripti
===================================
Bu script, test oturumlarını kolayca xlsx dosyasına kaydetmenizi sağlar.
Tester vaka çözme chatini kullandıktan sonra bu script ile test sonuçlarını kaydeder.

Kullanım:
    python qa_logger.py

Komutlar:
    yeni    → Yeni test kaydı oluştur
    liste   → Kayıtlı testleri listele
    hal     → Halüsinasyon detay kaydı ekle
    ozet    → Test özeti göster
    cikis   → Çıkış

Gereksinimler:
    pip install openpyxl
"""

import os
import sys
from datetime import datetime

import openpyxl

# ─── Sabit Değerler ───

XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'auto_fix_qa_tracker.xlsx')

SCENARIOS = {
    "S01": {"ad": "Araç Çalışmıyor (Akü)",       "arac": "2002 Toyota Corolla 1.6, Benzin",       "zorluk": "Kolay", "gercek_ariza": "Ölü akü (ömrünü tamamlamış)"},
    "S02": {"ad": "Marş Tık Sesi (Starter)",       "arac": "2006 Ford Focus 1.6 TDCi, Dizel",       "zorluk": "Kolay", "gercek_ariza": "Arızalı marş motoru"},
    "S03": {"ad": "Sol Far Yanmıyor (Sigorta)",    "arac": "2010 VW Polo 1.4, Benzin, Otomatik",     "zorluk": "Kolay", "gercek_ariza": "Sol far sigortası yanık"},
    "S04": {"ad": "Silecek Çalışmıyor (Motor)",    "arac": "2008 Hyundai Accent 1.5 CRDi, Dizel",    "zorluk": "Kolay", "gercek_ariza": "Silecek motoru arızalı"},
    "S05": {"ad": "Klima Soğutmuyor (Gaz Kaçağı)", "arac": "2012 Renault Clio 1.5 dCi, Dizel",       "zorluk": "Kolay", "gercek_ariza": "Klima gazı kaçağı (sistem boş)"},
    "S06": {"ad": "Aşırı Isınma (Termostat)",      "arac": "2005 Opel Astra 1.6, Benzin",            "zorluk": "Orta",  "gercek_ariza": "Termostat kapalı pozisyonda takılı"},
    "S07": {"ad": "Rölanti Titremesi (Buji)",      "arac": "1998 Fiat Palio 1.6, LPG",               "zorluk": "Orta",  "gercek_ariza": "3. silindir bujisi kirli/arızalı"},
    "S08": {"ad": "Sert Vites (ATF Sıvısı)",       "arac": "2011 Honda Civic 1.8 i-VTEC, Otomatik",  "zorluk": "Orta",  "gercek_ariza": "ATF sıvısı bozulmuş/yanmış"},
    "S09": {"ad": "Fren Sesi (Balata Aşınması)",   "arac": "2014 Kia Ceed 1.6 GDI, Benzin",          "zorluk": "Orta",  "gercek_ariza": "Ön balata aşınma göstergesine kadar bitmiş"},
    "S10": {"ad": "Check Engine (O2 Sensör)",      "arac": "2009 Peugeot 308 1.6 HDi, Dizel",        "zorluk": "Orta",  "gercek_ariza": "Downstream O2 sensörü arızalı"},
    "S11": {"ad": "Yağ Yakma (Piston Segman)",     "arac": "2003 BMW 320i E46, Benzin, Otomatik",     "zorluk": "Zor",   "gercek_ariza": "2. ve 3. silindir piston segmanları aşınmış"},
    "S12": {"ad": "Conta Patlaması (Coolant)",     "arac": "2007 VW Passat 1.9 TDI, Dizel",          "zorluk": "Zor",   "gercek_ariza": "Silindir kapak contası patlamış"},
    "S13": {"ad": "Su Pompası Arızası",            "arac": "2013 Renault Megane 1.5 dCi, Dizel",     "zorluk": "Zor",   "gercek_ariza": "Su pompası rulmanı bozuk + sızıntı"},
    "S14": {"ad": "Güç Kaybı (Triger Kayması)",    "arac": "2004 Fiat Doblo 1.9 JTD, Dizel",         "zorluk": "Zor",   "gercek_ariza": "Triger kayışı 2 diş atlamış"},
    "S15": {"ad": "LPG Arızası (ECU Kalibrasyon)", "arac": "2010 Hyundai Accent Era 1.5 CRDi, LPG",  "zorluk": "Zor",   "gercek_ariza": "LPG ECU haritası bozuk"},
}

SONUC_SECENEKLER = ["PASS", "FAIL", "PARTIAL"]
EVET_HAYIR = ["Evet", "Hayır"]
HALUCINASYON_TIPLER = ["Yok", "Uydurma Parça", "Yanlış Teşhis", "Protokol Dışı", "Dil Karışımı", "Diğer"]
KATMAN_SECENEKLER = ["Giriş Filtresi", "Prompt Hardening", "Çıkış Filtresi", "Tespit Edilemedi"]
AKSIYON_SECENEKLER = ["İnceleniyor", "Prompt Güncellendi", "Lambda Güncellendi", "Kapatıldı", "Yok Sayıldı"]


# ─── Yardımcı Fonksiyonlar ───

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')


def print_header():
    print()
    print("=" * 60)
    print("  🔧 Auto Fix AI — QA Test Kayıt Sistemi")
    print("=" * 60)
    print()


def print_scenarios():
    """Senaryo listesini göster."""
    print("\n📋 Mevcut Senaryolar:")
    print("-" * 75)
    print(f"  {'ID':<5} {'Zorluk':<8} {'Senaryo Adı':<35} {'Araç'}")
    print("-" * 75)

    current_difficulty = None
    for sid, info in SCENARIOS.items():
        if info['zorluk'] != current_difficulty:
            current_difficulty = info['zorluk']
            if sid != "S01":
                print()
        print(f"  {sid:<5} {info['zorluk']:<8} {info['ad']:<35} {info['arac'][:30]}")
    print("-" * 75)


def choose_from_list(prompt, options, allow_empty=False):
    """Listeden seçim yaptır."""
    for i, opt in enumerate(options, 1):
        print(f"  {i}. {opt}")

    while True:
        choice = input(f"\n{prompt} (1-{len(options)}): ").strip()
        if allow_empty and choice == "":
            return None
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(options):
                return options[idx]
        except ValueError:
            pass
        print("  ⚠️  Geçersiz seçim, tekrar deneyin.")


def get_input(prompt, required=True, default=None):
    """Kullanıcıdan input al."""
    suffix = f" [{default}]" if default else ""
    while True:
        val = input(f"{prompt}{suffix}: ").strip()
        if not val and default:
            return default
        if val or not required:
            return val
        print("  ⚠️  Bu alan zorunludur.")


def get_int_input(prompt, min_val=None, max_val=None, required=True):
    """Sayısal input al."""
    while True:
        val = input(f"{prompt}: ").strip()
        if not val and not required:
            return None
        try:
            num = int(val)
            if min_val is not None and num < min_val:
                print(f"  ⚠️  En az {min_val} olmalı.")
                continue
            if max_val is not None and num > max_val:
                print(f"  ⚠️  En fazla {max_val} olmalı.")
                continue
            return num
        except ValueError:
            print("  ⚠️  Lütfen bir sayı girin.")


def find_next_empty_row(ws, col_check=2, start_row=5, max_row=54):
    """Test Log'da sonraki boş satırı bul."""
    for row in range(start_row, max_row + 1):
        if ws.cell(row=row, column=col_check).value is None:
            return row
    return None


# ─── Ana Komutlar ───

def cmd_yeni_test():
    """Yeni test kaydı oluştur."""
    print("\n🆕 Yeni Test Kaydı")
    print("=" * 40)

    # 1. Senaryo seç
    print_scenarios()
    while True:
        sid = input("\n🎯 Senaryo ID (S01-S15): ").strip().upper()
        if sid in SCENARIOS:
            break
        print("  ⚠️  Geçersiz ID. S01-S15 arasında girin.")

    scenario = SCENARIOS[sid]
    print(f"\n  ✅ Seçilen: {scenario['ad']}")
    print(f"     Araç: {scenario['arac']}")
    print(f"     Zorluk: {scenario['zorluk']}")
    print(f"     Gerçek Arıza: {scenario['gercek_ariza']}")

    # 2. Tester adı
    tester_adi = get_input("\n👤 Tester Adı")

    # 3. Sonuç
    print(f"\n📊 Test Sonucu:")
    sonuc = choose_from_list("Sonuç", SONUC_SECENEKLER)

    # 4. Mesaj sayısı
    mesaj_sayisi = get_int_input("\n💬 Toplam Mesaj Sayısı (AI ile kaç mesaj yazıldı)", min_val=1)

    # 5. İpucu kullanıldı mı?
    print(f"\n💡 İpucu (Danışman Usta) Kullanıldı mı?")
    ipucu = choose_from_list("İpucu", EVET_HAYIR)

    # 6. Halüsinasyon var mı?
    print(f"\n🤖 Halüsinasyon Tespit Edildi mi?")
    print("  (AI gerçek olmayan bir arıza mı uydurdu? Sağlam parçayı arızalı mı gösterdi?)")
    halucinasyon = choose_from_list("Halüsinasyon", EVET_HAYIR)

    halucinasyon_tipi = "Yok"
    yanlis_parca = ""
    if halucinasyon == "Evet":
        print(f"\n🔍 Halüsinasyon Tipi:")
        halucinasyon_tipi = choose_from_list("Tip", HALUCINASYON_TIPLER[1:])  # "Yok" hariç

        yanlis_parca = get_input("  Yanlış gösterilen parça/bilgi (kısa açıklama)", required=False) or ""

    # 7. Prompt injection denendi mi?
    print(f"\n🛡️ Prompt Injection Denendi mi?")
    print("  (Oyuncunun AI'ı kandırmaya çalışıp çalışmadığı)")
    prompt_injection = choose_from_list("Injection", EVET_HAYIR)

    # 8. Yanıt dili doğru mu?
    print(f"\n🌐 AI Yanıt Dili Doğru mu? (Türkçe mi yanıt verdi?)")
    dil_dogru = choose_from_list("Dil", EVET_HAYIR)

    # 9. Süre
    sure = get_int_input("\n⏱️ Testin Süresi (dakika)", min_val=1, required=False)

    # 10. Güven skoru
    guven = get_int_input("\n⭐ Güven Skoru (1=çok kötü, 5=mükemmel)", min_val=1, max_val=5)

    # 11. Notlar
    notlar = get_input("\n📝 Ek Notlar (opsiyonel)", required=False) or ""

    # ─── XLSX'e yaz ───
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['Test Log']

    row = find_next_empty_row(ws)
    if row is None:
        print("\n  ❌ HATA: Test Log'da boş satır kalmadı! (Maks 50 test)")
        return

    tarih = datetime.now().strftime("%Y-%m-%d %H:%M")

    # A: Test ID (formül var, dokunma)
    ws.cell(row=row, column=2).value = tarih                    # B: Tarih
    ws.cell(row=row, column=3).value = tester_adi               # C: Tester Adı
    ws.cell(row=row, column=4).value = sid                      # D: Senaryo ID
    # E: Senaryo Adı (formül var / VLOOKUP)
    # F: Araç / Arıza (formül var / VLOOKUP)
    # G: Zorluk (formül var)
    ws.cell(row=row, column=8).value = sonuc                    # H: Sonuç
    ws.cell(row=row, column=9).value = mesaj_sayisi             # I: Mesaj Sayısı
    ws.cell(row=row, column=10).value = ipucu                   # J: İpucu Kullanıldı?
    ws.cell(row=row, column=11).value = halucinasyon            # K: Halüsinasyon?
    ws.cell(row=row, column=12).value = halucinasyon_tipi       # L: Halüsinasyon Tipi
    ws.cell(row=row, column=13).value = yanlis_parca            # M: Yanlış Parça
    ws.cell(row=row, column=14).value = prompt_injection        # N: Prompt Injection?
    ws.cell(row=row, column=15).value = dil_dogru               # O: Yanıt Dili Doğru?
    ws.cell(row=row, column=16).value = sure                    # P: Süre
    ws.cell(row=row, column=17).value = guven                   # Q: Güven Skoru
    ws.cell(row=row, column=18).value = notlar                  # R: Notlar
    # S: Flag (formül var, dokunma)

    wb.save(XLSX_PATH)
    test_no = row - 4
    print(f"\n  ✅ Test #{test_no} başarıyla kaydedildi! (Satır {row})")

    # Halüsinasyon varsa otomatik detay kaydı iste
    if halucinasyon == "Evet":
        print(f"\n  ⚠️  Halüsinasyon tespit edildi! Detay kaydı da eklenmeli.")
        cevap = input("  📝 Şimdi halüsinasyon detayı eklemek ister misin? (e/h): ").strip().lower()
        if cevap == 'e':
            cmd_halucinasyon_detay(prefill={
                "test_id": test_no,
                "senaryo_id": sid,
                "tarih": tarih,
                "tester": tester_adi,
                "tip": halucinasyon_tipi,
            })


def cmd_halucinasyon_detay(prefill=None):
    """Halüsinasyon detay kaydı ekle."""
    print("\n🔬 Halüsinasyon Detay Kaydı")
    print("=" * 40)

    if prefill:
        test_id = prefill["test_id"]
        senaryo_id = prefill["senaryo_id"]
        tarih = prefill["tarih"]
        tester = prefill["tester"]
        hal_tip = prefill["tip"]
        print(f"  (Test #{test_id}, {senaryo_id} için otomatik dolduruldu)")
    else:
        test_id = get_int_input("Test ID (Test Log'daki sıra numarası)", min_val=1)
        senaryo_id = get_input("Senaryo ID (S01-S15)").upper()
        tarih = datetime.now().strftime("%Y-%m-%d %H:%M")
        tester = get_input("Tester Adı")
        print(f"\n🔍 Halüsinasyon Tipi:")
        hal_tip = choose_from_list("Tip", HALUCINASYON_TIPLER[1:])

    # Detay bilgileri
    ai_yanit = get_input("\n🤖 AI'ın verdiği yanlış/halüsinasyon yanıtı (özet)")
    dogru_cevap = get_input("✅ Doğru cevap ne olmalıydı?")
    tetikleyen = get_input("💬 Hangi mesaj bunu tetikledi? (tester ne yazdı?)")

    print(f"\n🛡️ Hangi güvenlik katmanında sorun var?")
    katman = choose_from_list("Katman", KATMAN_SECENEKLER)

    print(f"\n📌 Aksiyon:")
    aksiyon = choose_from_list("Aksiyon", AKSIYON_SECENEKLER)

    # XLSX'e yaz
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['Halüsinasyon Detay']

    row = find_next_empty_row(ws, col_check=1, start_row=5)
    if row is None:
        print("\n  ❌ HATA: Halüsinasyon Detay'da boş satır kalmadı!")
        return

    ws.cell(row=row, column=1).value = test_id          # A: Test ID
    ws.cell(row=row, column=2).value = senaryo_id       # B: Senaryo ID
    ws.cell(row=row, column=3).value = tarih            # C: Tarih
    ws.cell(row=row, column=4).value = tester           # D: Tester
    ws.cell(row=row, column=5).value = hal_tip          # E: Halüsinasyon Tipi
    ws.cell(row=row, column=6).value = ai_yanit         # F: AI Yanıtı (özet)
    ws.cell(row=row, column=7).value = dogru_cevap      # G: Doğru Cevap
    ws.cell(row=row, column=8).value = tetikleyen       # H: Tetikleyen Mesaj
    ws.cell(row=row, column=9).value = katman           # I: Katman
    ws.cell(row=row, column=10).value = aksiyon         # J: Aksiyon

    wb.save(XLSX_PATH)
    print(f"\n  ✅ Halüsinasyon detayı kaydedildi! (Satır {row})")


def cmd_liste():
    """Kayıtlı testleri listele."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['Test Log']

    print("\n📋 Kayıtlı Testler")
    print("=" * 100)
    print(f"  {'#':<4} {'Tarih':<17} {'Tester':<12} {'Senaryo':<6} {'Sonuç':<8} {'Hal?':<5} {'Flag'}")
    print("-" * 100)

    count = 0
    for row in range(5, 55):
        tarih = ws.cell(row=row, column=2).value
        if tarih is None:
            break

        test_no = row - 4
        tester = ws.cell(row=row, column=3).value or "-"
        senaryo = ws.cell(row=row, column=4).value or "-"
        sonuc = ws.cell(row=row, column=8).value or "-"
        hal = ws.cell(row=row, column=11).value or "-"
        flag = "🚩" if (hal == "Evet" or sonuc == "FAIL") else ""

        print(f"  {test_no:<4} {str(tarih):<17} {tester:<12} {senaryo:<6} {sonuc:<8} {hal:<5} {flag}")
        count += 1

    if count == 0:
        print("  (Henüz test kaydı yok)")
    else:
        print(f"\n  Toplam: {count} test")


def cmd_ozet():
    """Test özeti göster."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['Test Log']

    total = 0
    pass_count = 0
    fail_count = 0
    partial_count = 0
    hal_count = 0
    injection_count = 0
    senaryo_test_count = {}

    for row in range(5, 55):
        tarih = ws.cell(row=row, column=2).value
        if tarih is None:
            break

        total += 1
        sonuc = ws.cell(row=row, column=8).value
        hal = ws.cell(row=row, column=11).value
        injection = ws.cell(row=row, column=14).value
        senaryo = ws.cell(row=row, column=4).value

        if sonuc == "PASS":
            pass_count += 1
        elif sonuc == "FAIL":
            fail_count += 1
        elif sonuc == "PARTIAL":
            partial_count += 1

        if hal == "Evet":
            hal_count += 1
        if injection == "Evet":
            injection_count += 1

        if senaryo:
            senaryo_test_count[senaryo] = senaryo_test_count.get(senaryo, 0) + 1

    print("\n📊 Test Özeti")
    print("=" * 50)
    print(f"  Toplam Test:         {total}")
    print(f"  ✅ PASS:             {pass_count}")
    print(f"  ❌ FAIL:             {fail_count}")
    print(f"  ⚠️  PARTIAL:         {partial_count}")
    print(f"  🤖 Halüsinasyon:     {hal_count}")
    print(f"  🛡️  Prompt Injection: {injection_count}")

    if total > 0:
        basari = pass_count / total * 100
        hal_oran = hal_count / total * 100
        print(f"\n  Başarı Oranı:        {basari:.1f}%")
        print(f"  Halüsinasyon Oranı:  {hal_oran:.1f}%")

    # Test edilmemiş senaryolar
    tested = set(senaryo_test_count.keys())
    all_scenarios = set(SCENARIOS.keys())
    untested = all_scenarios - tested

    if untested:
        print(f"\n  ⚠️  Test Edilmemiş Senaryolar ({len(untested)}):")
        for sid in sorted(untested):
            print(f"     {sid}: {SCENARIOS[sid]['ad']}")
    else:
        print(f"\n  ✅ Tüm senaryolar en az 1 kez test edildi!")


# ─── Ana Döngü ───

def main():
    if not os.path.exists(XLSX_PATH):
        print(f"❌ HATA: {XLSX_PATH} bulunamadı!")
        print("Önce update_xlsx.py scriptini çalıştırın.")
        sys.exit(1)

    clear_screen()
    print_header()

    while True:
        print("\n┌─────────────────────────────────┐")
        print("│  Komutlar:                      │")
        print("│    yeni  → Yeni test kaydı       │")
        print("│    liste → Kayıtlı testler       │")
        print("│    hal   → Halüsinasyon detay     │")
        print("│    ozet  → Test özeti             │")
        print("│    cikis → Çıkış                  │")
        print("└─────────────────────────────────┘")

        cmd = input("\n🔧 Komut: ").strip().lower()

        if cmd == "yeni":
            cmd_yeni_test()
        elif cmd == "liste":
            cmd_liste()
        elif cmd == "hal":
            cmd_halucinasyon_detay()
        elif cmd == "ozet":
            cmd_ozet()
        elif cmd in ("cikis", "çıkış", "exit", "q"):
            print("\n👋 Görüşmek üzere!")
            break
        else:
            print("  ⚠️  Geçersiz komut. yeni / liste / hal / ozet / cikis yazın.")


if __name__ == '__main__':
    main()
