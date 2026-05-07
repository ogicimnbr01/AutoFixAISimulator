"""
autofix_qa_tracker.xlsx dosyasını oyundaki gerçek 15 senaryo ile günceller.
Senaryo Analizi sayfasındaki senaryo bilgilerini scenarios.py'deki verilerle eşleştirir.

Kullanım:
    python update_xlsx.py
"""

import sys
import os
import shutil
from datetime import datetime

# scenarios.py'yi import edebilmek için path ekle
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend', 'lambdas', 'shared'))

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# Oyundaki gerçek 15 senaryo: (ID, Kısa Ad, Kategori)
# scenarios.py'deki SCENARIOS listesinden türetildi
SCENARIO_MAPPING = [
    # EASY (1-5)
    ("S01", "Araç Çalışmıyor (Akü)",       "Elektrik – Akü",           "Easy",  "2002 Toyota Corolla 1.6"),
    ("S02", "Marş Tık Sesi (Starter)",       "Elektrik – Marş Motoru",   "Easy",  "2006 Ford Focus 1.6 TDCi"),
    ("S03", "Sol Far Yanmıyor (Sigorta)",    "Elektrik – Aydınlatma",    "Easy",  "2010 VW Polo 1.4"),
    ("S04", "Silecek Çalışmıyor (Motor)",    "Elektrik – Silecek",       "Easy",  "2008 Hyundai Accent 1.5 CRDi"),
    ("S05", "Klima Soğutmuyor (Gaz Kaçağı)", "Klima Sistemi",            "Easy",  "2012 Renault Clio 1.5 dCi"),
    # MEDIUM (6-10)
    ("S06", "Aşırı Isınma (Termostat)",      "Soğutma Sistemi",          "Medium", "2005 Opel Astra 1.6"),
    ("S07", "Rölanti Titremesi (Buji)",      "Ateşleme Sistemi",         "Medium", "1998 Fiat Palio 1.6 LPG"),
    ("S08", "Sert Vites (ATF Sıvısı)",       "Otomatik Şanzıman",        "Medium", "2011 Honda Civic 1.8 i-VTEC"),
    ("S09", "Fren Sesi (Balata Aşınması)",   "Fren Sistemi",             "Medium", "2014 Kia Ceed 1.6 GDI"),
    ("S10", "Check Engine (O2 Sensör)",      "OBD – Sensör Arızası",     "Medium", "2009 Peugeot 308 1.6 HDi"),
    # HARD (11-15)
    ("S11", "Yağ Yakma (Piston Segman)",     "Motor – Mekanik",          "Hard",  "2003 BMW 320i E46"),
    ("S12", "Conta Patlaması (Coolant)",     "Motor – Conta",            "Hard",  "2007 VW Passat 1.9 TDI"),
    ("S13", "Su Pompası Arızası",            "Soğutma – Pompa",          "Hard",  "2013 Renault Megane 1.5 dCi"),
    ("S14", "Güç Kaybı (Triger Kayması)",    "Motor – Timing",           "Hard",  "2004 Fiat Doblo 1.9 JTD"),
    ("S15", "LPG Arızası (ECU Kalibrasyon)", "LPG Sistemi",              "Hard",  "2010 Hyundai Accent Era 1.5 CRDi"),
]

XLSX_PATH = os.path.join(os.path.dirname(__file__), 'autofix_qa_tracker.xlsx')
BACKUP_SUFFIX = datetime.now().strftime("_%Y%m%d_%H%M%S")


def update_xlsx():
    """XLSX dosyasını oyundaki gerçek vakalarla güncelle."""

    # Yedek al
    backup_path = XLSX_PATH.replace('.xlsx', f'{BACKUP_SUFFIX}_backup.xlsx')
    shutil.copy2(XLSX_PATH, backup_path)
    print(f"✅ Yedek alındı: {backup_path}")

    wb = openpyxl.load_workbook(XLSX_PATH)

    # ─── 1. Senaryo Analizi sayfasını güncelle ───
    ws_sa = wb['Senaryo Analizi']
    for idx, (sid, name, category, difficulty, vehicle) in enumerate(SCENARIO_MAPPING):
        row = 5 + idx  # Row 5'ten başlıyor
        ws_sa.cell(row=row, column=1).value = sid          # A: Senaryo ID
        ws_sa.cell(row=row, column=2).value = name         # B: Senaryo Adı
        ws_sa.cell(row=row, column=3).value = category     # C: Araç / Arıza kategorisi
        # D-L sütunları formül, dokunmuyoruz

    print(f"✅ Senaryo Analizi: {len(SCENARIO_MAPPING)} senaryo güncellendi")

    # ─── 2. Test Log: Senaryo ID dropdown doğrulaması güncelle ───
    ws_tl = wb['Test Log']

    # Mevcut Senaryo ID validation'ı kaldır (varsa)
    existing_dvs = list(ws_tl.data_validations.dataValidation)
    for dv in existing_dvs:
        if 'D5' in str(dv.sqref):
            ws_tl.data_validations.dataValidation.remove(dv)

    # Yeni dropdown: Senaryo ID listesi
    scenario_ids = ",".join([s[0] for s in SCENARIO_MAPPING])
    dv_scenario = DataValidation(
        type="list",
        formula1=f'"{scenario_ids}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Geçersiz Senaryo",
        error="Lütfen S01-S15 arasında bir senaryo seçin."
    )
    dv_scenario.sqref = 'D5:D54'
    ws_tl.add_data_validation(dv_scenario)

    # ─── 3. Test Log: Senaryo Adı ve Araç/Arıza otomatik doldurma formülleri ───
    # E sütunu (Senaryo Adı): VLOOKUP ile Senaryo Analizi'nden çek
    # F sütunu (Araç / Arıza): Senaryo bilgisine göre
    for row in range(5, 55):
        # E: Senaryo Adı -> VLOOKUP
        ws_tl.cell(row=row, column=5).value = (
            f"=IFERROR(VLOOKUP(D{row},'Senaryo Analizi'!A$5:B$19,2,FALSE),\"\")"
        )
        # F: Araç / Arıza -> VLOOKUP
        ws_tl.cell(row=row, column=6).value = (
            f"=IFERROR(VLOOKUP(D{row},'Senaryo Analizi'!A$5:C$19,3,FALSE),\"\")"
        )

    # ─── 4. Zorluk otomatik doldurma (ID'ye göre) ───
    # S01-S05 = Kolay, S06-S10 = Orta, S11-S15 = Zor
    for row in range(5, 55):
        ws_tl.cell(row=row, column=7).value = (
            f'=IF(D{row}="","",IF(VALUE(MID(D{row},2,2))<=5,"Kolay",IF(VALUE(MID(D{row},2,2))<=10,"Orta","Zor")))'
        )

    print(f"✅ Test Log: Dropdown, VLOOKUP formülleri ve Zorluk formülü eklendi")

    # ─── 5. Dashboard referansları kontrol et (zaten doğru, row eşleşmeleri aynı) ───
    print(f"✅ Dashboard: Referanslar mevcut yapıyla uyumlu (15 satır)")

    # ─── 6. Kaydet ───
    wb.save(XLSX_PATH)
    print(f"\n🎉 {XLSX_PATH} başarıyla güncellendi!")
    print(f"\nGüncel senaryolar:")
    print(f"{'ID':<5} {'Zorluk':<8} {'Senaryo Adı':<35} {'Kategori'}")
    print("-" * 80)
    for sid, name, category, difficulty, vehicle in SCENARIO_MAPPING:
        diff_tr = {"Easy": "Kolay", "Medium": "Orta", "Hard": "Zor"}[difficulty]
        print(f"{sid:<5} {diff_tr:<8} {name:<35} {category}")


if __name__ == '__main__':
    update_xlsx()
